import openpyxl
import pandas as pd
import glob
import numpy as np

import_file_path = 'C:\\Users\\81805\\Desktop\\py\\T.xlsx' #ファイルのパスの指定
excel_sheet_name = 'FeedbackLight' #編集するシート名
export_file_path = 'C:\\Users\\81805\\Desktop\\py\\output.xlsx' #出力するファイルのパス

l_high = 255  #照射光強度の強い値，弱い値を設定
l_low = 0

df_raster = pd.read_excel(import_file_path, sheet_name = excel_sheet_name)

book = openpyxl.load_workbook(import_file_path, data_only=True) #各セルの値を数値として読み取る（ブック変数）
sheet = book[excel_sheet_name] #シートをシート変数に格納
Ncolumn = book[excel_sheet_name].max_column - 1 #ブックの行数を変数Nrowに格納
Nrow = book[excel_sheet_name].max_row #ブックの行数を変数Nrowに格納

oscc = np.ones(Ncolumn)
M = 10000 #最大の発振回数
osclist = np.zeros((M,Ncolumn)) #excelに出力する為のリストを生成する

for j in range(0,Ncolumn,1):
    k = 0 #kの値のリセット
    for i in range(3, Nrow, 1): #行数分だけ繰り返すループ中のiの値は参照している行
        if 255 == sheet.cell(row = i, column = j+2).value and 0 == sheet.cell(row = i+1, column = j+2).value:#
            osclist[k][j] = sheet.cell(row = i+1, column = 1).value #リストに発振時刻を
            k = k + 1
            oscc[j] = oscc[j] + 1

book = openpyxl.Workbook() #新しいExcelファイルを作成
sheet = book['Sheet'] #シート関数を新たに定義
book.save(export_file_path) #空のexcelファイルを保存

sheet['A1'] = '発振時刻'
sheet['B1'] = 'ラスタープロット用'

oscc = [int(s) for s in oscc] #整数型に変換
for j in range(0,Ncolumn,1):
    for i in range(1,oscc[j],1):
        sheet.cell(row = i*3+1, column = 2*j+1).value = osclist[i][j]
        sheet.cell(row = i*3+1, column = 2*j+2).value = Ncolumn - j + 1
        sheet.cell(row = i*3+2, column = 2*j+1).value = osclist[i][j]
        sheet.cell(row = i*3+2, column = 2*j+2).value = Ncolumn - j

book.save(export_file_path)
